import os
import sys

import csv
import glob
import shutil
import openpyxl

from datetime import datetime

from config.settings import *

from modules.garbagecollector import cleanup
from modules.sharepoint365 import Sharepoint365


class Linux:
    def __init__(self, os_type, change_request, patch_cycle):
        self.server_details = []
        self.os_type = os_type
        self.change_request = change_request
        self.patch_list_name = "%s Compliance Patch List" % patch_cycle

    def get_server_data(self):
        csv_files = glob.glob('./' + '*.csv')
        
        # format csv files to get needed details
        for csv_file in csv_files:
            with open(csv_file,"r") as file:
                no_header = file.read().split('\n', 1)[1]
                formatted_file = no_header.split(',1\n')

                for data in formatted_file:
                    details = data.replace('\n', ' ').split(',')
                    
                    if len(details) == 12:
                        self.server_details.append([
                            details[0],
                            self.patch_list_name,
                            "Compliant",
                            self.os_type,
                            details[8],
                            details[9],
                            details[10],
                            details[11],
                            "1",
                            self.change_request
                        ])

    def generate_csv(self):
        now = datetime.now()
        date_time = now.strftime("%Y%m%dT%H%M%S")
        csv_file_name = '%s - Compliance Report-%s.csv' % (self.change_request, date_time)
        header = [
            "Computer Name",
            "Patch List Name",
            "Compliance Status",
            "OS Type",
            "CAH - Patch Tags",
            "Custom Tags",
            "Operating System",
            "Cloud Provider",
            "Count",
            "Change Number"
        ]

        with open(csv_file_name, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(header)
            writer.writerows(self.server_details)

        # create output folder
        output_dir = "output"
        try:
            os.mkdir(output_dir)
        except FileExistsError:
            pass

        shutil.move(csv_file_name, output_dir)

    def run(self):
        self.get_server_data()
        cleanup(self.change_request)
        self.generate_csv()


class Windows:
    def __init__(self, os_type, change_request, patch_cycle):
        self.server_details = []
        self.os_type = os_type
        self.change_request = change_request
        self.patch_cycle = patch_cycle

    def ask_question(self, question):
        answer = input(question)
        return answer

    def get_compliance_tracker(self):
        # get all xlsx files
        workbooks = glob.glob('./' + '*.xlsx')
        question = "Excel Files:\n"
        for id, workbook in enumerate(workbooks):
            question += "(%s) - %s\n" % (id + 1, workbook)

        question += "\nChoose file: "

        answer = self.ask_question(question)

        try:
            answer = int(answer) - 1
        except ValueError:
            printer("Not a valid option.")
            sys.exit()

        if answer < 0:
            print("Not a valid option")
            sys.exit()

        if answer >= len(workbooks):
            print("Not a valid option")
            sys.exit()

        return workbooks[answer]

    def download_compliance_tracker(self):
        sharepoint_api = Sharepoint365()

        # get compliance tracker to be downloaded
        source_directory = "%s/%s" % (SHAREPOINT_DOC_URL, self.patch_cycle)
        files = sharepoint_api.list_files(source_directory)
        file_name = None

        for file in files:
            if "compliance tracker" in file.name.lower():
                file_name = file.name
                break

        # download target file
        sharepoint_api.download_file(source_directory, file_name)

        return file_name

    def get_active_spreadsheet(self):
        file_name = self.download_compliance_tracker()
        workbook = openpyxl.load_workbook(file_name)
        question = "%s Spreadsheets:\n" % workbook
        sheet_id = 1
        
        for sheetname in workbook.sheetnames:
            question += "%s - %s\n" % (sheet_id, sheetname)
            sheet_id += 1

        question += "\nChoose Spreadsheets: "
        
        answer = self.ask_question(question)

        try:
            answer = int(answer) - 1
        except ValueError:
            printer("Not a valid option.")
            sys.exit()

        if answer < 0:
            print("Not a valid option")
            sys.exit()

        if answer >= len(workbook.sheetnames):
            print("Not a valid option")
            sys.exit()

        return workbook[workbook.sheetnames[answer]]

    def get_server_data(self):
        # get sheetname
        spreadsheet = self.get_active_spreadsheet()
        index = 0
        cn_index = None
        cpt_index = None
        ct_index = None
        os_index = None
        cp_index = None
        filter_index = None
        
        for header in spreadsheet.iter_cols(min_row=1, max_row=1, values_only=True):
            if header[0] is None:
                continue

            if header[0].lower() == "Computer Name".lower():
                cn_index = index
            elif header[0].lower() == "CAH - Patch Tags".lower():
                cpt_index = index
            elif header[0].lower() == "Custom Tags".lower():
                ct_index = index
            elif header[0].lower() == "Operating System".lower():
                os_index = index
            elif header[0].lower() == "Cloud Provider - CAH".lower():
                cp_index = index
            elif "kb sensor" in header[0].lower():
                filter_index = index

            index += 1
        
        for details in spreadsheet.iter_rows(min_row=2, values_only=True):
            if all(elem is None for elem in details):
                continue

            if details[filter_index].lower() != "compliant":
                continue
            
            patch_list_name = "%s Compliance Patch List" % self.patch_cycle
            self.server_details.append([
                    details[cn_index],
                    patch_list_name,
                    "Compliant",
                    self.os_type,
                    details[cpt_index],
                    details[ct_index],
                    details[os_index],
                    details[cp_index],
                    "1",
                    self.change_request
                ])

    def generate_csv(self):
        now = datetime.now()
        date_time = now.strftime("%Y%m%dT%H%M%S")
        csv_file_name = '%s - Compliance Report-%s.csv' % (self.change_request, date_time)
        header = [
            "Computer Name",
            "Patch List Name",
            "Compliance Status",
            "OS Type",
            "CAH - Patch Tags",
            "Custom Tags",
            "Operating System",
            "Cloud Provider",
            "Count",
            "Change Number"
        ]

        with open(csv_file_name, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(header)
            writer.writerows(self.server_details)

        # create output folder
        output_dir = "output"
        try:
            os.mkdir(output_dir)
        except FileExistsError:
            pass

        shutil.move(csv_file_name, output_dir)

    def run(self):
        self.get_server_data()
        self.generate_csv()
        